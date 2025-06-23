import streamlit as st
import time
import json
import re
import openai
import io
import os
from dotenv import load_dotenv
from pypdf import PdfReader
from docx import Document
from openpyxl import load_workbook

load_dotenv() # Charge les variables du fichier .env dans l'environnement

OPENAI_API_KEY = os.getenv("OPENAI_API_KEY") or st.secrets.get("OPENAI_API_KEY")
PERPLEXITY_API_KEY = os.getenv("PERPLEXITY_API_KEY") or st.secrets.get("PERPLEXITY_API_KEY")

# Nettoyage
if OPENAI_API_KEY:
    OPENAI_API_KEY = OPENAI_API_KEY.strip()
if PERPLEXITY_API_KEY:
    PERPLEXITY_API_KEY = PERPLEXITY_API_KEY.strip()

# --- Configuration & Constantes ---

st.set_page_config(layout="wide")

REQUIRED_DOCS_LIST = [
    "Formulaire de demande / questionnaire dûment rempli",
    "Liste détaillée des véhicules de la flotte (Excel ou structuré)",
    "Historique de sinistralité sur les 3 à 5 dernières années",
    "Relevé d'informations / Attestation du précédent assureur",
    "Extrait Kbis récent de l'entreprise",
    "RIB de l'entreprise",
    "Conditions particulières souhaitées",
    "Document non identifiable"
]

# --- Récupération des Clés API depuis l'environnement ---
# Les clés sont maintenant chargées depuis le fichier .env


# --- Fonctions d'Extraction de Texte ---

def extract_text_from_file(uploaded_file):
    """Extrait le texte de différents types de fichiers, en gérant les onglets multiples pour Excel."""
    try:
        if uploaded_file.type == "application/pdf":
            pdf_reader = PdfReader(io.BytesIO(uploaded_file.getvalue()))
            text = ""
            for page in pdf_reader.pages:
                text += page.extract_text() or ""
            return text
        elif uploaded_file.type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
            doc = Document(io.BytesIO(uploaded_file.getvalue()))
            return "\n".join([para.text for para in doc.paragraphs])
        elif uploaded_file.type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
            workbook = load_workbook(filename=io.BytesIO(uploaded_file.getvalue()))
            full_text = ""
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                full_text += f"--- DEBUT CONTENU DE L'ONGLET: '{sheet_name}' ---\n"
                rows_data = []
                for row in sheet.iter_rows(values_only=True):
                    if any(cell is not None for cell in row):
                        row_str = " | ".join([str(cell) if cell is not None else "" for cell in row])
                        rows_data.append(row_str)
                full_text += "\n".join(rows_data)
                full_text += f"\n--- FIN CONTENU DE L'ONGLET: '{sheet_name}' ---\n\n"
            return full_text
        elif "text" in uploaded_file.type:
            return uploaded_file.getvalue().decode("utf-8")
        else:
            return None
    except Exception as e:
        st.warning(f"Impossible de lire le fichier '{uploaded_file.name}': {e}")
        return None

# --- Fonctions des Agents ---

def identify_documents_in_content_with_llm(filename, content_snippet, client):
    """Utilise l'IA pour identifier un ou plusieurs documents à partir du contenu d'un fichier."""
    
    prompt = f"""
    Vous êtes un assistant expert en souscription d'assurance flotte automobile.
    Votre tâche est d'analyser le contenu d'un fichier pour déterminer quels documents requis il contient. Un seul fichier peut contenir plusieurs types de documents (par exemple, un fichier Excel avec plusieurs onglets).

    Voici la liste des types de documents possibles que nous recherchons :
    <document_types>
    {json.dumps(REQUIRED_DOCS_LIST, indent=2, ensure_ascii=False)}
    </document_types>

    Voici le nom du fichier et son contenu (qui peut contenir plusieurs sections/onglets) :
    <filename>{filename}</filename>
    <content>
    {content_snippet[:8000]}
    </content>

    Analysez le contenu et déterminez TOUS les types de documents de la liste ci-dessus qui sont présents dans ce fichier.
    
    Retournez votre réponse exclusivement au format JSON, avec une seule clé "documents_identifies". La valeur de cette clé doit être une LISTE de chaînes de caractères correspondant aux types de documents trouvés.
    Si le fichier contient une liste de véhicules ET un historique de sinistres, la liste doit contenir ces deux éléments.
    Si le fichier ne correspond à aucun type de document ou est illisible, retournez une liste vide.

    Exemple de réponse pour un fichier Excel contenant des véhicules et des sinistres:
    {{
      "documents_identifies": [
        "Liste détaillée des véhicules de la flotte (Excel ou structuré)",
        "Historique de sinistralité sur les 3 à 5 dernières années"
      ]
    }}
    
    Exemple si le document est incompréhensible:
    {{
      "documents_identifies": []
    }}
    """
    try:
        response = client.chat.completions.create(
            model="gpt-4o", # Utilisation d'un modèle plus puissant pour cette tâche complexe
            messages=[
                {"role": "system", "content": "Vous êtes un expert en assurance qui identifie les documents contenus dans des fichiers."},
                {"role": "user", "content": prompt}
            ],
            response_format={"type": "json_object"},
            temperature=0.0,
        )
        result_json = response.choices[0].message.content
        analysis_result = json.loads(result_json)
        return analysis_result.get("documents_identifies", [])
    except Exception as e:
        st.error(f"Erreur lors de l'identification du fichier {filename}: {e}")
        return []

def extract_key_information_with_llm(all_content_text, client):
    """Utilise l'IA pour extraire les informations clés de l'ensemble des documents."""

    prompt = f"""
    Vous êtes un expert en souscription d'assurance qui analyse un dossier de demande de devis complet.
    Voici le contenu de tous les fichiers fournis, concaténés en un seul texte :
    <dossier_complet>
    {all_content_text[:15000]}
    </dossier_complet>

    Votre tâche est de lire attentivement l'intégralité du dossier et d'extraire les informations suivantes.
    Retournez votre réponse exclusivement au format JSON. Si une information n'est pas trouvée, mettez la valeur `null` ou une liste vide [].

    1.  **"nom_entreprise"**: Le nom légal de l'entreprise.
    2.  **"secteur_activite"**: Le secteur d'activité de l'entreprise.
    3.  **"region"**: La région ou le département principal de l'entreprise.
    4.  **"nombre_vehicules"**: Le nombre total de véhicules dans la flotte. Instruction : comptez les lignes du tableau de véhicules.
    5.  **"usage_flotte"**: L'usage principal de la flotte.
    6.  **"type_flotte"**: Le type de véhicules majoritaire.
    7.  **"chiffre_affaires_annuel"**: Le dernier chiffre d'affaires annuel.
    8.  **"historique_sinistralite_resume"**: Un résumé court de l'historique de sinistralité.
    9.  **"garanties_souhaitees"**: Une liste des garanties demandées (objets JSON avec "garantie", "incluse", "franchise_eur").
    10. **"liste_vehicules"**: La liste détaillée des véhicules. Chaque véhicule doit être un objet JSON. Extrayez les colonnes telles que "marque", "modele", "immatriculation", "date_mise_circulation", "valeur", etc.

    Exemple de format JSON de sortie attendu :
    {{
      "nom_entreprise": "Transport Express SARL",
      "secteur_activite": "Transport routier de marchandises",
      "region": "Île-de-France",
      "nombre_vehicules": 2,
      "usage_flotte": "Transport national de marchandises",
      "type_flotte": "Camions de livraison",
      "chiffre_affaires_annuel": "2.5 M€",
      "historique_sinistralite_resume": "3 sinistres responsables sur les 36 derniers mois",
      "garanties_souhaitees": [
        {{"garantie": "Responsabilité civile", "incluse": "Oui", "franchise_eur": 500}}
      ],
      "liste_vehicules": [
        {{ "immatriculation": "AA-123-BB", "marque_modele": "Renault Master", "valeur": 25000 }},
        {{ "immatriculation": "CC-456-DD", "marque_modele": "Peugeot Expert", "valeur": 22000 }}
      ]
    }}
    """
    try:
        with st.spinner("Extraction des données en cours..."):
            response = client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {"role": "system", "content": "Vous êtes un expert en extraction de données d'assurance au format JSON."},
                    {"role": "user", "content": prompt}
                ],
                response_format={"type": "json_object"},
                temperature=0.0,
            )
            result_json = response.choices[0].message.content
            return json.loads(result_json)
    except Exception as e:
        st.error(f"Une erreur est survenue lors de l'extraction des informations clés : {e}")
        return None

def smart_intake_agent(uploaded_files, openai_client):
    """
    L'agent Smart Intake analyse le contenu de chaque fichier, vérifie la complétude,
    puis extrait les informations clés.
    """
    st.write("🤖 **Agent Smart Intake en action...**")
    
    all_identified_doc_types = []
    files_content = {}

    with st.spinner("Analyse du contenu de tous les documents en cours... Cela peut prendre un moment."):
        for file in uploaded_files:
            content = extract_text_from_file(file)
            files_content[file.name] = content
            
            if content is None or not content.strip():
                continue
            
            doc_types_found = identify_documents_in_content_with_llm(file.name, content, openai_client)
            
            if doc_types_found:
                all_identified_doc_types.extend(doc_types_found)

    st.write("---")
    st.write("### ✅ Bilan de complétude du dossier")
    
    present_docs = sorted(list(set(all_identified_doc_types)))
    all_required = [doc for doc in REQUIRED_DOCS_LIST if doc != "Document non identifiable"]
    missing_docs = sorted(list(set(all_required) - set(present_docs)))
    
    st.write("#### Documents Fournis (consolidés sur tous les fichiers) :")
    if present_docs:
        for doc in present_docs:
            st.markdown(f"- <span style='color:green;'>{doc}</span>", unsafe_allow_html=True)
    else:
        st.write("_Aucun document requis n'a été identifié._")

    if missing_docs:
        st.write("#### 🔻 Documents Manquants :")
        for doc in missing_docs:
            st.markdown(f"- <span style='color:red;'>{doc}</span>", unsafe_allow_html=True)
        st.error("Le dossier est incomplet. Veuillez fournir les documents manquants pour continuer.")
        return False, None
    
    st.success("Dossier jugé complet par l'IA ! Poursuite du traitement.")
    st.markdown("---")

    # Nouvelle étape: Extraction des informations clés
    full_content_string = "\n\n".join([f"--- DEBUT FICHIER: {name} ---\n{content}" for name, content in files_content.items() if content])
    extracted_data = extract_key_information_with_llm(full_content_string, openai_client)

    if extracted_data:
        st.write("### 📝 Informations Clés Extraites par l'IA")
        
        # Afficher le tableau des véhicules en premier
        if 'liste_vehicules' in extracted_data and extracted_data['liste_vehicules']:
            st.write("#### Détail de la Flotte de Véhicules")
            st.dataframe(extracted_data['liste_vehicules'], use_container_width=True)
        
        # Afficher les garanties dans un tableau
        if 'garanties_souhaitees' in extracted_data and extracted_data['garanties_souhaitees']:
            st.write("#### Garanties Demandées")
            st.dataframe(extracted_data['garanties_souhaitees'], use_container_width=True)
        
        # Afficher le reste des informations
        st.write("#### Autres Informations")
        other_info = {k: v for k, v in extracted_data.items() if k not in ['garanties_souhaitees', 'liste_vehicules']}
        st.json(other_info)
        
        # Mapper les clés extraites vers le format attendu par les agents suivants
        final_data_for_agents = {
            "Nom de l'entreprise": extracted_data.get("nom_entreprise"),
            "Secteur d'activité": extracted_data.get("secteur_activite"),
            "Région": extracted_data.get("region"),
            "Nombre de véhicules": extracted_data.get("nombre_vehicules"),
            "Usage": extracted_data.get("usage_flotte"),
            "Type de flotte": extracted_data.get("type_flotte"),
            "Chiffre d'affaires": extracted_data.get("chiffre_affaires_annuel"),
            "Historique de sinistralité": extracted_data.get("historique_sinistralite_resume"),
            "Liste des véhicules": extracted_data.get("liste_vehicules", [])
        }
        return True, final_data_for_agents
    else:
        st.error("L'extraction des informations clés a échoué.")
        return False, None

def enrichment_layer_agent(data, perplexity_client, openai_client):
    """
    Agent qui utilise Perplexity pour la recherche web et OpenAI pour l'extraction.
    """
    st.write("🤖 **Agent Enrichment Layer en action...**")
    
    # 1. Définir les questions de recherche
    company_name = data["Nom de l'entreprise"]
    activity_sector = data["Secteur d'activité"]
    region = data["Région"]
    
    search_queries = {
        "sector_claim_rate": f"Quel est le taux de sinistralité moyen dans le secteur d'activité '{activity_sector}' en France ?",
        "geo_risk": f"Quels sont les risques de vol, vandalisme et d'accident pour les véhicules d'entreprise dans la région '{region}' en France ?",
        "telematics_risk_score_info": f"Comment un score de risque télématique influence-t-il l'assurance pour une flotte de '{data['Type de flotte']}' ?"
    }
    
    # 2. Effectuer les recherches avec Perplexity
    search_results = {}

    try:
        for key, query in search_queries.items():
            with st.spinner(f"Recherche en cours : {query}"):
                response = perplexity_client.chat.completions.create(
                    model="llama-3.1-sonar-small-128k-online",
                    messages=[
                        {"role": "system", "content": "Vous êtes un assistant de recherche. Fournissez des réponses factuelles et concises basées sur les informations disponibles sur Internet."},
                        {"role": "user", "content": query},
                    ],
                )
                search_results[key] = response.choices[0].message.content
    except openai.AuthenticationError:
        st.error("Erreur d'authentification Perplexity. Veuillez vérifier votre clé API.")
        return data
    except Exception as e:
        st.error(f"Une erreur est survenue lors de la recherche Perplexity : {e}")
        return data

    with st.expander("Voir les résultats bruts de la recherche"):
        st.json(search_results)
        
    # 3. Extraire les informations structurées avec OpenAI
    
    extraction_prompt = f"""
    Vous êtes un expert en analyse de données pour l'assurance.
    Voici des informations brutes provenant d'une recherche sur Internet :
    <search_results>
    {json.dumps(search_results, indent=2, ensure_ascii=False)}
    </search_results>

    Votre tâche est d'extraire les informations clés suivantes et de les retourner dans un format JSON strict.
    Si une information n'est pas clairement trouvable, mettez "Non trouvé".

    Format JSON attendu :
    {{
        "taux_sinistralite_secteur": "Ex: 12%",
        "analyse_risque_geo": "Un résumé court des risques de la région.",
        "facteur_risque_telematique": "Un résumé court de l'influence de la télématique."
    }}

    Ne retournez que le JSON.
    """
    
    try:
        with st.spinner("Extraction des données structurées..."):
            response = openai_client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[
                    {"role": "system", "content": "Vous êtes un expert en extraction de données JSON."},
                    {"role": "user", "content": extraction_prompt}
                ],
                response_format={"type": "json_object"},
                temperature=0.0,
            )
            extracted_info_json = response.choices[0].message.content
            extracted_info = json.loads(extracted_info_json)
    except Exception as e:
        st.error(f"Une erreur est survenue lors de l'extraction par OpenAI : {e}")
        return data

    # 4. Fusionner les données
    enriched_data = data.copy()
    new_keys = [
        "Taux de sinistralité du secteur", 
        "Analyse du risque géographique", 
        "Info sur le score télématique"
    ]
    enriched_data[new_keys[0]] = extracted_info.get("taux_sinistralite_secteur", "Non trouvé")
    enriched_data[new_keys[1]] = extracted_info.get("analyse_risque_geo", "Non trouvé")
    enriched_data[new_keys[2]] = extracted_info.get("facteur_risque_telematique", "Non trouvé")
    
    st.write("Données enrichies par l'IA :")
    st.json({k: v for k, v in enriched_data.items() if k in new_keys})
    return enriched_data

def rule_engine_agent(data):
    """
    Simulates the Rule Engine agent.
    - Helps with underwriting analysis.
    - Returns a JSON (in French) with all information for the quoting system.
    """
    st.write("🤖 **Agent Rule Engine en action...**")
    with st.spinner("Analyse du dossier pour la souscription et génération du JSON..."):
        time.sleep(2)
        
        # Simulate underwriting rules
        decision = "Favorable"
        commentaires = "Le profil de l'entreprise est bon. Les données enrichies par l'IA confirment un risque modéré pour le secteur et la géographie."
        
        st.write("Analyse de souscription :")
        st.info(f"**Décision :** {decision}\n\n**Commentaires :** {commentaires}")

        # Generate JSON for the quoting system (in French)
        quote_system_json = {
            "informations_client": {
                "nom_entreprise": data.get("Nom de l'entreprise"),
                "siren": data.get("SIREN", "N/A"), # Enrichi plus tard
                "sante_financiere": data.get("Santé financière (fictif)", "N/A") # Enrichi plus tard
            },
            "informations_flotte": {
                "nombre_vehicules": data.get("Nombre de véhicules"),
                "type_flotte": data.get("Type de flotte"),
                "usage": data.get("Usage"),
                "liste_vehicules": data.get("Liste des véhicules", [])
            },
            "analyse_risque": {
                "historique_sinistralite": data.get("Historique de sinistralité"),
                "taux_sinistralite_secteur": data.get("Taux de sinistralité du secteur"),
                "risque_geographique": data.get("Analyse du risque géographique"),
                "info_score_telematique": data.get("Info sur le score télématique"),
                "decision_souscription": decision,
                "commentaire_souscription": commentaires
            },
            "parametres_tarification": {
                "niveau_risque": "Moyen",
                "segment": "Transport Logistique"
            }
        }
        
        st.write("JSON (en français) pour le système de tarification :")
        st.code(json.dumps(quote_system_json, indent=4, ensure_ascii=False), language="json")
        return quote_system_json


# --- Interface Principale ---

st.title("Automatisation du Traitement des Devis Flotte Auto")

st.header("Étape 1: Réception et Analyse IA du Dossier")

with st.expander("Voir la liste des documents requis", expanded=False):
    st.markdown("Pour que le dossier soit considéré comme complet, veuillez fournir les documents suivants :")
    for doc in REQUIRED_DOCS_LIST:
        st.write(f"- {doc}")

uploaded_files = st.file_uploader(
    "Veuillez charger tous les documents du dossier de demande de devis.",
    type=['pdf', 'xlsx', 'docx', 'csv', 'txt'],
    accept_multiple_files=True
)

if uploaded_files:
    if st.button("Lancer l'analyse du dossier", type="primary"):
        # La vérification se fait maintenant sur les variables chargées depuis l'environnement
        if not OPENAI_API_KEY or not PERPLEXITY_API_KEY:
            st.error("🛑 Clés API non trouvées. Assurez-vous d'avoir un fichier .env correctement configuré, ou si l'application est déployée, que les secrets sont bien configurés dans Streamlit Cloud.")
        else:
            # Initialisation des clients, en forçant la conversion en chaîne pour plus de robustesse
            openai_client = openai.OpenAI(api_key=str(OPENAI_API_KEY))
            perplexity_client = openai.OpenAI(api_key=str(PERPLEXITY_API_KEY), base_url="https://api.perplexity.ai")

            # --- Smart Intake ---
            is_complete, extracted_data = smart_intake_agent(uploaded_files, openai_client)
            
            # --- Processus conditionnel ---
            if is_complete:
                st.markdown("---")
                
                # --- Enrichment Layer ---
                st.header("Étape 2: Enrichment Layer")
                enriched_data = enrichment_layer_agent(extracted_data, perplexity_client, openai_client)
                
                st.markdown("---")

                # --- Rule Engine ---
                st.header("Étape 3: Rule Engine & Souscription")
                quote_json = rule_engine_agent(enriched_data)

                # Prépare le JSON pour le téléchargement
                json_string_to_download = json.dumps(quote_json, indent=4, ensure_ascii=False)
                
                # Crée un nom de fichier sûr
                company_name_safe = quote_json.get("informations_client", {}).get("nom_entreprise", "client").replace(" ", "_")

                st.markdown("---")

                col1, col2 = st.columns(2)

                with col1:
                    if st.button("✅ Envoyer au tarificateur", type="primary", use_container_width=True):
                        with st.spinner("Connexion au système de tarification..."):
                            time.sleep(2)
                        st.success("Les données ont été envoyées avec succès au système de tarification !")
                        st.balloons()
                
                with col2:
                    st.download_button(
                       label="📥 Télécharger le JSON",
                       data=json_string_to_download,
                       file_name=f"donnees_tarification_{company_name_safe}.json",
                       mime="application/json",
                       use_container_width=True
                    )

                st.balloons()
                st.success("Processus de traitement du devis terminé !") 